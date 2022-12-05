import csv
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image
import flask
from flask import request, Flask, redirect, url_for
from werkzeug.utils import secure_filename

UPLOAD_FOLDER = "."
app = Flask(__name__)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

@app.route("/success/<name>")
def success(name):
    return "Succesfully Generated Mark Sheets!!" % name

def fun_consice(positive_marks, negative_marks):
    def add_column_in_csv(input_file, output_file, transform_row, dict_statusAns):
        with open(input_file, "r") as read_obj, open(
            output_file, "w", newline=""
        ) as write_obj:
            
            csv_reader = csv.reader(read_obj)
            csv_writer = csv.writer(write_obj)
            for row in csv_reader:
                row.append(dict_statusAns[row[6]])
                transform_row(row, csv_reader.line_num)
                csv_writer.writerow(row)

    location = "marksheet"

    def consice_marksheet():
        if not os.path.exists(location):
            os.makedirs(location)
        dict_scoreafter_neg = {}
        dict_statusAns = {}
        with open("responses.csv", "r") as file:
            reader = csv.reader(file)
            for row in reader:
                no_of_corrects = 0
                no_of_wrongs = 0
                no_of_unattempted = 0
                score_after_negative = 0
                for i in range(7, len(answer_list)):
                    if row[i] == answer_list[i]:
                        no_of_corrects += 1
                    elif row[i] == "":
                        no_of_unattempted += 1
                    else:
                        no_of_wrongs += 1
                total_marks = (
                    no_of_corrects + no_of_wrongs + no_of_unattempted
                ) * positive_marks
                score_after_negative = (no_of_corrects * positive_marks) + (
                    no_of_wrongs * negative_marks
                )
                dict_scoreafter_neg[row[6]] = (
                    str(score_after_negative) + "/" + str(total_marks)
                )
                dict_statusAns[row[6]] = (
                    "["
                    + str(no_of_corrects)
                    + ","
                    + str(no_of_wrongs)
                    + ","
                    + str(no_of_unattempted)
                    + "]"
                )

        add_column_in_csv(
            "responses.csv",
            file_path,
            lambda row, line_num: row.insert(6, dict_scoreafter_neg[row[6]]),
            dict_statusAns,
        )

    file_name = "consice_marksheet" + ".csv"
    file_path = os.path.join(location, file_name)
    dict_master_roll_name = {}
    master_roll = []
    abs_roll = []
    with open("master_roll.csv", "r") as file:
        reader1 = csv.reader(file)
        for row in reader1:
            master_roll.append(row[0])
            dict_master_roll_name[row[0]] = row[1]
    with open("responses.csv", "r") as file:
        reader = csv.reader(file)
        answer_found = 0
        roll_list_responses = []
        answer_list = []
        for row in reader:
            if row[6] == "ANSWER":
                answer_found = 1
                answer_list = row
            roll_list_responses.append(row[6])
    if answer_found == 1:
        consice_marksheet()
        for roll1 in master_roll:
            if roll1 not in roll_list_responses:
                abs_roll.append(roll1)
        with open(file_path, "a") as file:
            csv_writer = csv.writer(file)
            for rollno in abs_roll:
                csv_writer.writerow(
                    [
                        "",
                        "",
                        "ABSENT",
                        dict_master_roll_name[rollno],
                        "",
                        "",
                        "ABSENT",
                        rollno,
                    ]
                )
    else:
        print("no roll number with ANSWER is present, Cannot Process!")


def fun_rollwise(positive_marks, negative_marks):
    location = "marksheet"
    def score_generator():
        dict_r_w_n = {}
        dict_scoreafter_neg = {}
        dict_marking = {}
        dict_total = {}
        dict_name_options = {}
        with open("responses.csv", "r") as file:
            reader = csv.reader(file)
            for row in reader:
                no_of_corrects = 0
                no_of_wrongs = 0
                no_of_unattempted = 0
                score_after_negative = 0
                for i in range(7, len(answer_list)):
                    if row[i] == answer_list[i]:
                        no_of_corrects += 1
                    elif row[i] == "":
                        no_of_unattempted += 1
                    else:
                        no_of_wrongs += 1
                total_marks = (
                    no_of_corrects + no_of_wrongs + no_of_unattempted
                ) * positive_marks
                score_after_negative = (no_of_corrects * positive_marks) + (
                    no_of_wrongs * negative_marks
                )
                dict_scoreafter_neg[row[6]] = (
                    str(score_after_negative) + "/" + str(total_marks)
                )
                dict_r_w_n[row[6]] = [
                    no_of_corrects,
                    no_of_wrongs,
                    no_of_unattempted,
                    no_of_corrects + no_of_wrongs + no_of_unattempted,
                ]
                dict_marking[row[6]] = [positive_marks, negative_marks]
                dict_total[row[6]] = [
                    no_of_corrects * positive_marks,
                    no_of_wrongs * negative_marks,
                    dict_scoreafter_neg[row[6]],
                ]
                dict_name_options[row[6]] = row
        return dict_r_w_n, dict_marking, dict_total, dict_name_options

    dict_master_roll_name = {}
    master_roll = []
    with open("master_roll.csv", "r") as file:
        reader1 = csv.reader(file)
        for row in reader1:
            master_roll.append(row[0])
            dict_master_roll_name[row[0]] = row[1]
    with open("responses.csv", "r") as file:
        reader = csv.reader(file)
        answer_found = 0
        roll_list_responses = []
        answer_list = []
        for row in reader:
            if row[6] == "ANSWER":
                answer_found = 1
                answer_list = row
            roll_list_responses.append(row[6])
    if answer_found == 1:
        dict_r_w_n, dict_marking, dict_total, dict_name_options = score_generator()
        for roll in roll_list_responses:
            book = openpyxl.Workbook()
            sheet = book.active
            bold_style = Font(name="Century", size=12, bold=True)
            normal_style = Font(name="century", size=12)
            blue_style = Font(name="Century", size=12, color="0000FF")
            Green_style = Font(name="Century", size=12, color="008000")
            Red_style = Font(name="Century", size=12, color="FF0000")
            bd = openpyxl.styles.Side(style='thin', color="000000")
            highlight = openpyxl.styles.Border(left=bd, top=bd, right=bd, bottom=bd)
            sheet.cell(row=6, column=1).value = "Name:"
            sheet.cell(row=6, column=1).font = normal_style
            sheet.cell(row=6, column=2).value = dict_name_options[roll][3]
            sheet.cell(row=6, column=2).font = bold_style
            sheet.cell(row=7, column=1).value = "Roll Numer:"
            sheet.cell(row=7, column=1).font = normal_style
            sheet.cell(row=7, column=2).value = roll
            sheet.cell(row=7, column=2).font = bold_style
            sheet.cell(row=6, column=4).value = "Exam:"
            sheet.cell(row=6, column=4).font = normal_style
            sheet.cell(row=6, column=5).value = "quiz"
            sheet.cell(row=6, column=5).font = bold_style
            fieldnames = [
                "Right",
                "Wrong",
                "Not Attempt",
                "Max",
            ]
            for col_no, fieldname in enumerate(fieldnames, start=2):
                sheet.cell(row=9, column=col_no).value = fieldname
                sheet.cell(row=9, column=col_no).font = bold_style
            fieldnames0 = [
                "No.",
                "Marking",
                "Total",
            ]
            for row_no, fieldname in enumerate(fieldnames0, start=10):
                sheet.cell(row=row_no, column=1).value = fieldname
                sheet.cell(row=row_no, column=1).font = bold_style
            fieldnames1 = [
                dict_r_w_n[roll][0],
                dict_marking[roll][0],
                dict_total[roll][0],
            ]
            for row_no, fieldname in enumerate(fieldnames1, start=10):
                sheet.cell(row=row_no, column=2).value = fieldname
                sheet.cell(row=row_no, column=2).font = Green_style
            fieldnames2 = [
                dict_r_w_n[roll][1],
                dict_marking[roll][1],
                dict_total[roll][1],
            ]
            for row_no, fieldname in enumerate(fieldnames2, start=10):
                sheet.cell(row=row_no, column=3).value = fieldname
                sheet.cell(row=row_no, column=3).font = Red_style
            sheet.cell(row=10, column=4).value = dict_r_w_n[roll][2]
            sheet.cell(row=10, column=4).font = normal_style
            sheet.cell(row=11, column=4).value = 0
            sheet.cell(row=11, column=4).font = normal_style
            sheet.cell(row=10, column=5).value = dict_r_w_n[roll][3]
            sheet.cell(row=10, column=5).font = normal_style
            sheet.cell(row=12, column=5).value = dict_total[roll][2]
            sheet.cell(row=12, column=5).font = blue_style
            sheet.cell(row=15, column=1).value = "Student Ans"
            sheet.cell(row=15, column=1).font = bold_style
            sheet.cell(row=15, column=2).value = "Correct Ans"
            sheet.cell(row=15, column=2).font = bold_style
            for i in range(0, len(answer_list) - 7):
                sheet.cell(row=i + 16, column=1).value = dict_name_options[roll][i + 7]
                sheet.cell(row=i + 16, column=2).value = answer_list[i + 7]
                sheet.cell(row=i + 16, column=2).font = blue_style
                if dict_name_options[roll][i + 7] == answer_list[i + 7]:
                    sheet.cell(row=i + 16, column=1).font = Green_style
                else:
                    sheet.cell(row=i + 16, column=1).font = Red_style
            for i in "ABCDE":
                sheet.column_dimensions[i].width = 18
            for col in sheet.iter_cols():
                for cell in col:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(row=6, column=1).alignment = Alignment(
                horizontal="right", vertical="center"
            )
            sheet.cell(row=7, column=1).alignment = Alignment(
                horizontal="right", vertical="center"
            )
            sheet.cell(row=6, column=4).alignment = Alignment(
                horizontal="right", vertical="center"
            )
            sheet.cell(row=6, column=2).alignment = Alignment(
                horizontal="left", vertical="center"
            )
            sheet.cell(row=7, column=2).alignment = Alignment(
                horizontal="left", vertical="center"
            )
            sheet.cell(row=6, column=5).alignment = Alignment(
                horizontal="left", vertical="center"
            )
            sheet.merge_cells("A5:E5")
            sheet.cell(row=5, column=1).value = "Mark Sheet"
            sheet.cell(row=5, column=1).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            Big_style = Font(name="Century", size=18, underline="single")
            sheet.cell(row=5, column=1).font = Big_style
            img = openpyxl.drawing.image.Image("IITP_logo.JPEG")
            img.height = 80
            img.width = 630
            sheet.add_image(img)
            for i in range(1,3):
                for j in range(15,len(answer_list)+9):
                    sheet.cell(row=j,column=i).border=highlight
            for i in range(9,13):
                for j in range(1,6):
                    sheet.cell(row=i,column=j).border=highlight
            if not os.path.exists(location):
                os.makedirs(location)
            file_name = roll + ".xlsx"
            file_path = os.path.join(location, file_name)
            book.save(file_path)
        for roll1 in master_roll:
            if roll1 not in roll_list_responses:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.cell(row=6, column=1).value = "Name:"
                ws.cell(row=6, column=1).font = normal_style
                ws.cell(row=6, column=2).value = dict_master_roll_name[roll1]
                ws.cell(row=6, column=2).font = bold_style
                ws.cell(row=7, column=1).value = "Roll Numer:"
                ws.cell(row=7, column=1).font = normal_style
                ws.cell(row=7, column=2).value = roll1
                ws.cell(row=7, column=2).font = bold_style
                ws.cell(row=6, column=4).value = "Exam:"
                ws.cell(row=6, column=4).font = normal_style
                ws.cell(row=6, column=5).value = "quiz"
                ws.cell(row=6, column=5).font = bold_style
                for i in "ABCDE":
                    ws.column_dimensions[i].width = 18
                ws.merge_cells("A5:E5")
                ws.cell(row=5, column=1).value = "Mark Sheet"
                ws.cell(row=5, column=1).alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                Big_style = Font(name="Century", size=18, bold=True)
                ws.cell(row=5, column=1).font = Big_style
                img = openpyxl.drawing.image.Image("IITP_logo.JPEG")
                img.height = 80
                img.width = 630
                ws.add_image(img)
                file_name1 = roll1 + ".xlsx"
                file_path1 = os.path.join(location, file_name1)
                wb.save(file_path1)
    else:
        print("no roll number with ANSWER is present, Cannot Process!")


@app.route("/login", methods=["POST", "GET"])
def login():
    if request.method == "POST":
        files = flask.request.files.getlist("file")
        for file in files:
            if file.filename == "master_roll.csv" or file.filename == "responses.csv":
                file.save(os.path.join(app.config["UPLOAD_FOLDER"], file.filename))

        user1 = request.form["nm1"]
        user2 = request.form["nm2"]
        positive_marks = float(user1)
        negative_marks = float(user2)
        print("submit_button value: ", request.form["submit_button"])
        if request.form["submit_button"] == "Generate Consise Marksheet":
            fun_consice(positive_marks, negative_marks)
        elif request.form["submit_button"] == "Generate Roll Number wise Marksheet":
            fun_rollwise(positive_marks, negative_marks)
            print("success")
        return redirect(url_for("success", name=user1))
    else:
        user = request.args.get("nm1")
        return redirect(url_for("success", name=user))

if __name__ == "__main__":
    app.run(debug=True)
